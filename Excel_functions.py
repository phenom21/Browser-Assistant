from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import openpyxl
def getrowcount(path, sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return(sheet.max_row)
def getcolumncount(path, sheetname):
    workbook=openpyxl.load_workbook(path)
    sheet=workbook.get_sheet_by_name(sheetname)
    return(sheet.max_column)
def readata(path, sheetname, rown, columnn):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    return(sheet.cell(row=rown, column=columnn).value)
def writedata(path, sheetname, rown, columnn, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rown, column=columnn).value=data
    workbook.save(path)
